#
# AutoHABR Deluxe™️ (cooperatively yours to use and modify)
#
#
# Redesigned by Micah Carroll in 2023 (original author: Rafael Soto, date unbeknownst)
#
# Description: This script automates the process of merging the latest expenses report (report.xlsx)
#              into the current summary.xlsx. The output is saved to a different file.
#


import openpyxl

CONTACT_EMAIL = "mdc@berkeley.edu"


#######################
# GETTING USER INPUTS #
#######################


def parse_user_inputs(_summary):
    test = False

    print("What is the Report Date? Please input in format: MM/DD/YYYY")
    _report_date = input() if not test else "01/01/2020"
    # Validate the date is in the correct format
    assert len(_report_date.split("/")) == 3, "Date must be in format MM/DD/YYYY"

    # Changing the Report Date in the Ending Balances tab of summary.xlsx (it propagates to all other tabs)
    _summary["Ending Balances"].cell(row=2, column=5).value = _report_date

    print("What semester is this report for? (e.g. 'Spring 2023')")
    _semester = input() if not test else "Spring 2023"
    assert any(
        [s in _semester for s in ["Spring", "Fall", "Summer"]]
    ), "Semester must include 'Spring', 'Fall', or 'Summer'"

    # Changing the Days so Far.
    # This number will be used to:
    # 1. Calculate the percentage of the fiscal year spanned by `beginning of semester` to `input date`
    # 2. Calculate budgets for supplies
    # 3. Various other things
    print("How many days of the semester are we in so far?")
    daysin = int(input()) if not test else 1

    # Setting the Days so Far in the Budgets tab of summary.xlsx (to be inputted value)
    budgetssheet = _summary["Budgets"]
    budgetssheet.cell(row=2, column=2).value = daysin

    # Total days in the semester
    # TODO: 120? There's a couple of days off, 120*3=360 instead of 365
    totaldays = budgetssheet.cell(row=3, column=2).value

    # Calculating the percentage of the fiscal year the selected window of time (start of semester to report date)
    # corresponds to
    percentage_of_fiscal_year = round((daysin / totaldays), 2)

    # Setting the percentage of the fiscal year in the Budgets tab of summary.xlsx
    # (this will affect e.g. the Gas & Electric budget calcuations)
    budgetssheet.cell(row=29, column=2).value = percentage_of_fiscal_year
    return _semester, _report_date


################
# PARSE REPORT #
################


def is_expense_code(name):
    if name is None or name == "":
        return False
    return all(c.isdigit() for c in name[:4])


def parse_report(_report_house_names, _report):
    # Dictionary for storing the house accounts for each house
    _house_accounts_d = {}

    # Dictionary for storing the maintenance costs for each house
    _maintainance_d = {}

    # iterates through sheets in report excelworkbook
    # Looks through a page and records the key as the numerical account number
    # adds this houseaccount dictionary to allhouseaccounts dictionary
    for r_housename in _report_house_names:
        r_house_sheet = _report[r_housename]
        houseaccounts = {}
        house_maintenance_running_cost = 0
        for idx in range(1, r_house_sheet.max_row):
            expense_name = r_house_sheet.cell(row=idx, column=1).value
            expense_value = r_house_sheet.cell(row=idx, column=3).value

            if not is_expense_code(expense_name):
                continue

            expense_code = expense_name[:4]

            # Treat maintenance expenses separately
            if expense_code[:2] == "56":
                house_maintenance_running_cost += expense_value

            houseaccounts[expense_code] = expense_value

        _house_accounts_d[r_housename] = houseaccounts
        _maintainance_d[r_housename] = house_maintenance_running_cost
    return _house_accounts_d, _maintainance_d


##################
# UPDATE SUMMARY #
##################

# iterates through the summary excelworkbook and inputs expense/income
# for corresponding accounts number/row using dictionaries


def populate_s_house_sheet(_s_house_sheet, _house_account):
    for i in range(1, _s_house_sheet.max_row):

        expense_code = str(_s_house_sheet.cell(row=i, column=1).value)
        # Evaluate the expression in column 2 of the same row
        expense_name = _s_house_sheet.cell(row=i, column=2).internal_value

        if not is_expense_code(expense_code):
            continue

        expense_code = expense_code[:4]

        # The two relevant columns in the summary sheet (for budget/income entries and expenses)
        budget_income_cell = _s_house_sheet.cell(row=i, column=5)
        expense_cell = _s_house_sheet.cell(row=i, column=6)

        if expense_code == "5650":
            # Overwrite maintenance expenses with the running total for house
            expense_cell.value = maintenance_d[r_house_code]

        elif expense_code == "6440":
            # Parking is currently being ignored. Unclear if it could be automated too?
            pass

        elif expense_code == "2020":
            # House account rollover doesn't change, so it won't be in the report
            pass

        elif expense_code == "5820+5830":
            # NOTE: This part of the code might be dead. If this code is still here next time someone sees it, it's
            #  probably safe to delete it.
            assert False, f"Contact f{CONTACT_EMAIL} if you see this message."

            if "5820" in _house_account or "5830" in _house_account:
                h_acc_5820 = _house_account["5820"]
                h_acc_5830 = _house_account["5830"]

                if "5820" in _house_account and "5830" in _house_account:
                    if h_acc_5820 + h_acc_5830 > 0:
                        expense_cell.value = abs(h_acc_5820 + h_acc_5830)
                    else:
                        budget_income_cell.value = abs(h_acc_5820 + h_acc_5830)

                elif "5820" in _house_account and "5830" not in _house_account:
                    if h_acc_5820 > 0:
                        expense_cell.value = abs(h_acc_5820)
                    else:
                        budget_income_cell.value = abs(h_acc_5820)
                else:
                    if h_acc_5830 > 0:
                        expense_cell.value = abs(h_acc_5830)
                    else:
                        budget_income_cell.value = abs(h_acc_5830)
            else:
                raise ValueError(f"Something wen't wrong. Contact {CONTACT_EMAIL}")

        elif expense_code in _house_account:
            if _house_account[expense_code] > 0:
                expense_cell.value = _house_account[expense_code]
            else:
                budget_income_cell.value = abs(_house_account[expense_code])

        else:
            print(f"Expense code {expense_code} for `{expense_name}` not found in report.")


if __name__ == "__main__":
    print(
        f"\nWelcome to your premium AutoHABR Deluxe™️ experience! Sit back and relax, while my poor little CPU does the dirty work for you.\n"
    )
    print(
        f"If you have any ideas for improvement for the script, feel free to contact my developer at: {CONTACT_EMAIL}\n"
    )

    print("Now loading the report.xlsx and summary.xlsx files...\n")
    report = openpyxl.load_workbook("report.xlsx")
    summary = openpyxl.load_workbook("summary.xlsx")
    print("Loading completed.\n")

    # Tab names in report are house names in a specific code format
    r_house_codes = report.sheetnames
    num_houses = len(r_house_codes)

    # Tab names in summary are house names in a different code format, with some extra summary tabs
    # NOTE: assumption is that the order of houses in the report and summary is the same.
    s_house_codes = summary.sheetnames[:num_houses]
    assert len(s_house_codes) == len(r_house_codes), f"Number of houses in report and summary do not match."

    # Try catch for incorrectly formatted inputs
    try:
        # Get user inputs re: dates for the summary sheet
        semester, report_date = parse_user_inputs(summary)
    except ValueError as e:
        print("Inputs in incorrect format. Try running the script again. \n", e)
        exit(1)

    print("Parsing report...\n")
    # Parse the report and get the house accounts and maintenance costs
    house_accounts_d, maintenance_d = parse_report(r_house_codes, report)

    print("Populating summary sheet...\n")
    # Populate the summary sheet for all houses
    for s_house_code, r_house_code in zip(s_house_codes, r_house_codes):

        s_house_sheet = summary[s_house_code]
        house_account = house_accounts_d[r_house_code]
        print(f"\n\nPopulating summary sheet for house: {s_house_code} / {r_house_code}\n")

        populate_s_house_sheet(s_house_sheet, house_account)

    print(f"\n\nIf any of the expense codes should have been found but weren't, contact {CONTACT_EMAIL}\n")

    file_friendly_date = report_date.replace("/", ".")
    filename = f"HABR Summary {semester} {file_friendly_date}.xlsx"

    print(f"Saving HABR Summary as `{filename}` (sometimes this takes a while)...")
    summary.save(filename)
    print("...aaaand it's done! Hope you have a great day! :)")
